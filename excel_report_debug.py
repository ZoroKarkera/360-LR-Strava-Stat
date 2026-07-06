import os

from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from models import Activity, Athlete
from statistics import (
    get_current_week_start,
    get_heatmap,
    get_leaderboard,
    get_recent_runs,
    get_report_start_date,
    get_summary,
)


REPORT_DIR = "reports"
REPORT_FILE = "360_Long_Runners_Report.xlsx"
DAYS = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]

COLORS = {
    "navy": "1F4E78",
    "blue": "D9EAF7",
    "green": "70AD47",
    "green_light": "E2F0D9",
    "gold": "F4B183",
    "grey": "F3F6F8",
    "white": "FFFFFF",
    "text": "1F2933",
    "muted": "6B7280",
    "border": "D9E2EC",
}


def generate_excel(filename=None):
    """Generate the club Excel report and return the saved file path."""
    output_path = filename or os.path.join(REPORT_DIR, REPORT_FILE)
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "360 Long Runners"
    ws.sheet_view.showGridLines = False

    report_start = get_report_start_date()
    week_start = get_current_week_start()

    summary = get_summary(start_date=report_start)
    leaderboard = get_leaderboard(start_date=report_start)
    heatmap = get_heatmap(start_date=week_start)
    recent_runs = get_recent_runs(limit=20, start_date=report_start)
    achievements = get_achievements(start_date=report_start)
    date_range = get_date_range_label(report_start)

    write_title(ws, date_range)
    write_summary(ws, summary)
    write_leaderboard(ws, leaderboard)
    write_heatmap(ws, heatmap, week_start)
    write_achievements(ws, achievements)
    write_recent_runs(ws, recent_runs)
    apply_page_formatting(ws)

    wb.save(output_path)
    print(f"Excel generated: {output_path}")
    return output_path


def write_title(ws, date_range):
    ws.merge_cells("A1:H1")
    ws["A1"] = "360 Long Runners"
    ws["A1"].font = Font(size=22, bold=True, color=COLORS["white"])
    ws["A1"].fill = PatternFill("solid", fgColor=COLORS["navy"])
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 34

    ws.merge_cells("A2:H2")
    ws["A2"] = date_range
    ws["A2"].font = Font(size=11, italic=True, color=COLORS["muted"])
    ws["A2"].alignment = Alignment(horizontal="center")


def write_summary(ws, summary):
    section_header(ws, "A4", "Club Summary")

    cards = [
        ("Distance", f"{summary['distance']:.1f} km"),
        ("Runs", summary["runs"]),
        ("Members", summary["runners"]),
        ("Elevation", f"{summary['elevation']:.0f} m"),
        ("Avg HR", summary["avg_hr"] if summary["avg_hr"] else "-"),
    ]

    start_col = 1
    for index, (label, value) in enumerate(cards):
        col = start_col + index
        top = ws.cell(row=5, column=col)
        bottom = ws.cell(row=6, column=col)

        top.value = label
        top.font = Font(size=10, bold=True, color=COLORS["muted"])
        top.fill = PatternFill("solid", fgColor=COLORS["grey"])
        top.alignment = Alignment(horizontal="center")

        bottom.value = value
        bottom.font = Font(size=15, bold=True, color=COLORS["text"])
        bottom.fill = PatternFill("solid", fgColor=COLORS["grey"])
        bottom.alignment = Alignment(horizontal="center")

        for cell in (top, bottom):
            cell.border = thin_border()


def write_leaderboard(ws, leaderboard):
    section_header(ws, "A9", "Leaderboard Since 01-Jun")
    headers = ["Rank", "Runner", "Distance", "Runs"]
    write_header_row(ws, 10, 1, headers)

    if not leaderboard:
        write_empty_state(ws, 11, 1, "No leaderboard data yet.")
        return

    for row_offset, runner in enumerate(leaderboard, start=1):
        row = 10 + row_offset
        ws.cell(row=row, column=1).value = row_offset
        ws.cell(row=row, column=2).value = runner["runner"]
        ws.cell(row=row, column=3).value = runner["distance"]
        ws.cell(row=row, column=4).value = runner["runs"]
        ws.cell(row=row, column=3).number_format = '0.0 "km"'
        style_body_row(ws, row, 1, 4)


def write_heatmap(ws, heatmap, week_start):
    section_header(
        ws,
        "F9",
        f"Current Week Heatmap ({week_start.strftime('%d-%b')})",
    )

    headers = ["Runner"] + DAYS + ["Total"]
    write_header_row(ws, 10, 6, headers)

    # Get every club member, sorted alphabetically
    athletes = (
        Athlete.query
        .order_by(Athlete.firstname.asc())
        .all()
    )
    print("=== HEATMAP ATHLETES ===")
    print([a.firstname for a in athletes])
    if not athletes:
        write_empty_state(ws, 11, 6, "No members found.")
        return

    start_row = 11

    for row_offset, athlete in enumerate(athletes):

        row = start_row + row_offset
        runner = athlete.firstname

        ws.cell(row=row, column=6).value = runner

        total = 0

        for day_index, day in enumerate(DAYS, start=7):

            distance = round(
                heatmap.get(runner, {}).get(day, 0),
                1
            )

            total += distance

            cell = ws.cell(row=row, column=day_index)
            cell.value = distance
            cell.number_format = "0.0"

        total_cell = ws.cell(row=row, column=14)
        total_cell.value = round(total, 1)
        total_cell.number_format = '0.0 "km"'

        style_body_row(ws, row, 6, 14)

    last_row = start_row + len(athletes) - 1

    ws.conditional_formatting.add(
        f"G{start_row}:M{last_row}",
        ColorScaleRule(
            start_type="num",
            start_value=0,
            start_color="FFFFFF",
            mid_type="percentile",
            mid_value=50,
            mid_color=COLORS["green_light"],
            end_type="max",
            end_color=COLORS["green"],
        ),
    )


def write_achievements(ws, achievements):
    section_header(ws, "A28", "Achievements Since 01-Jun")
    headers = ["Metric", "Winner", "Value"]
    write_header_row(ws, 29, 1, headers)

    if not achievements:
        write_empty_state(ws, 30, 1, "No achievements yet.")
        return

    for row_offset, item in enumerate(achievements, start=1):
        row = 29 + row_offset
        ws.cell(row=row, column=1).value = item["metric"]
        ws.cell(row=row, column=2).value = item["winner"]
        ws.cell(row=row, column=3).value = item["value"]
        style_body_row(ws, row, 1, 3)


def write_recent_runs(ws, recent_runs):
    section_header(ws, "F28", "Recent Activities Since 01-Jun")
    headers = ["Date", "Runner", "Activity", "Distance", "Pace", "Elev", "HR"]
    write_header_row(ws, 29, 6, headers)

    if not recent_runs:
        write_empty_state(ws, 30, 6, "No recent runs yet.")
        return

    for row_offset, activity in enumerate(recent_runs, start=1):
        row = 29 + row_offset
        values = [
            activity["date"],
            activity["runner"],
            activity["activity"],
            activity["distance"],
            activity["pace"],
            activity["elevation"],
            round(activity["hr"], 1) if activity["hr"] else "",
        ]

        for col_offset, value in enumerate(values, start=6):
            ws.cell(row=row, column=col_offset).value = value

        ws.cell(row=row, column=9).number_format = '0.00 "km"'
        ws.cell(row=row, column=11).number_format = '0 "m"'
        style_body_row(ws, row, 6, 12)


def get_achievements(start_date=None):
    query = Activity.query
    if start_date is not None:
        query = query.filter(Activity.start_date >= start_date)

    activities = query.order_by(Activity.start_date.desc()).all()
    if not activities:
        return []

    athlete_names = {
        athlete.athlete_id: athlete.firstname
        for athlete in Athlete.query.all()
    }

    longest_run = max(activities, key=lambda item: item.distance or 0)
    highest_elevation = max(
        activities,
        key=lambda item: item.total_elevation_gain or 0,
    )

    hr_activities = [
        activity for activity in activities if activity.average_heartrate
    ]
    lowest_hr = (
        min(hr_activities, key=lambda item: item.average_heartrate)
        if hr_activities
        else None
    )

    leaderboard = get_leaderboard(start_date=start_date)
    most_runs = max(leaderboard, key=lambda item: item["runs"]) if leaderboard else None

    rows = [
        {
            "metric": "Longest Run",
            "winner": athlete_names.get(longest_run.athlete_id, "Unknown"),
            "value": f"{(longest_run.distance or 0) / 1000:.2f} km",
        },
        {
            "metric": "Most Runs",
            "winner": most_runs["runner"] if most_runs else "-",
            "value": f"{most_runs['runs']} runs" if most_runs else "-",
        },
        {
            "metric": "Highest Elevation",
            "winner": athlete_names.get(highest_elevation.athlete_id, "Unknown"),
            "value": f"{highest_elevation.total_elevation_gain or 0:.0f} m",
        },
    ]

    if lowest_hr:
        rows.append(
            {
                "metric": "Lowest Average HR",
                "winner": athlete_names.get(lowest_hr.athlete_id, "Unknown"),
                "value": f"{lowest_hr.average_heartrate:.1f} bpm",
            }
        )

    return rows


def get_date_range_label(report_start):
    last_activity = (
        Activity.query
        .filter(Activity.start_date >= report_start)
        .order_by(Activity.start_date.desc())
        .first()
    )

    start_label = report_start.strftime("%d-%b-%Y")

    if not last_activity:
        return f"{start_label} to today"

    return (
        f"{start_label} to "
        f"{last_activity.start_date.strftime('%d-%b-%Y')}"
    )


def section_header(ws, cell_ref, label):
    cell = ws[cell_ref]
    cell.value = label
    cell.font = Font(size=14, bold=True, color=COLORS["navy"])
    cell.alignment = Alignment(vertical="center")


def write_header_row(ws, row, start_col, headers):
    for offset, header in enumerate(headers):
        cell = ws.cell(row=row, column=start_col + offset)
        cell.value = header
        cell.font = Font(bold=True, color=COLORS["white"])
        cell.fill = PatternFill("solid", fgColor=COLORS["navy"])
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border()


def style_body_row(ws, row, start_col, end_col):
    fill = PatternFill("solid", fgColor=COLORS["white"] if row % 2 else COLORS["grey"])
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.border = thin_border()
        cell.alignment = Alignment(vertical="top", wrap_text=True)


def write_empty_state(ws, row, col, message):
    cell = ws.cell(row=row, column=col)
    cell.value = message
    cell.font = Font(italic=True, color=COLORS["muted"])


def thin_border():
    side = Side(style="thin", color=COLORS["border"])
    return Border(left=side, right=side, top=side, bottom=side)


def apply_page_formatting(ws):
    widths = {
        "A": 16,
        "B": 18,
        "C": 16,
        "D": 12,
        "E": 12,
        "F": 16,
        "G": 18,
        "H": 12,
        "I": 12,
        "J": 16,
        "K": 11,
        "L": 10,
        "M": 10,
        "N": 12,
    }

    for column, width in widths.items():
        ws.column_dimensions[column].width = width

    for row in range(1, ws.max_row + 1):
        ws.row_dimensions[row].height = 22

    for column_index in range(1, ws.max_column + 1):
        column_letter = get_column_letter(column_index)
        for cell in ws[column_letter]:
            cell.alignment = cell.alignment.copy(vertical="center")

    ws.freeze_panes = "A10"
    ws.auto_filter.ref = f"A10:N{ws.max_row}"
